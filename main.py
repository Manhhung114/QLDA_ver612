import os
import re
import sys
import sqlite3
from pathlib import Path
from datetime import date, datetime, timedelta

from PySide6.QtCore import Qt, QDate, QRectF, QPointF
from PySide6.QtGui import QColor, QBrush, QPen, QFont, QPolygonF
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QComboBox, QLineEdit, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QDialog, QFormLayout, QDialogButtonBox, QDateEdit,
    QSpinBox, QTextEdit, QFileDialog, QFrame, QSplitter, QGraphicsView,
    QGraphicsScene, QGraphicsRectItem, QGraphicsPolygonItem, QAbstractItemView, QTabWidget
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

from mpp_reader import MppComReader, MppReadError
from document_manager import DocumentManagerPage
from drawing_manager import DrawingManagerPage
from report_dashboard import ReportDashboardPage
from legal_manager import LegalDocumentsPage
from ai_manager import AIAssistantPage
from settings_manager import SettingsPage
from access_control import desktop_access_state

APP_DIR = Path(__file__).resolve().parent
DB_PATH = APP_DIR / "qlda_tiendo_v2.db"
DATE_FMT = "%Y-%m-%d"


def progress_delta(planned, actual):
    try:
        return int(round(float(actual))) - int(round(float(planned)))
    except Exception:
        return 0


def progress_evaluation(planned, actual, tolerance=1):
    delta = progress_delta(planned, actual)
    if delta > tolerance:
        return f"Nhanh +{delta}%"
    if delta < -tolerance:
        return f"Chậm {abs(delta)}%"
    return "Đúng KH"


def calc_progress_status(start_s, end_s, planned, actual, status_date=None):
    """Trạng thái tiến độ tại ngày báo cáo.

    Quy tắc V3.6: TT=100% luôn là Hoàn thành. Nếu chưa đạt 100% và đã
    vượt ngày Kết thúc thì là Chậm tiến độ. Số ngày trễ được tính riêng
    bằng calculate_delay_days().
    """
    status_date = status_date or date.today()
    try:
        actual = int(round(float(actual)))
        planned = int(round(float(planned)))
    except Exception:
        return "Chưa xác định"
    if actual >= 100:
        return "Hoàn thành"
    try:
        start = datetime.strptime(start_s, DATE_FMT).date()
        end = datetime.strptime(end_s, DATE_FMT).date()
    except Exception:
        return "Chưa xác định"
    if status_date < start and actual <= 0:
        return "Chưa bắt đầu"
    if status_date > end:
        return "Chậm tiến độ"
    delta = progress_delta(planned, actual)
    if delta < -1:
        return "Chậm tiến độ"
    if delta > 1:
        return "Nhanh tiến độ"
    return "Đúng tiến độ"


def calculate_delay_days(end_s, actual, actual_finish_date="", status_date=None):
    """Tính số ngày trễ so với ngày Kết thúc.

    - Chưa hoàn thành (<100%): trễ tăng theo ngày báo cáo/hiện tại.
    - Đã hoàn thành (100%): khóa số ngày trễ tại ngày lần đầu đạt 100%.
    - Dữ liệu 100% cũ chưa có ngày hoàn thành: không tự suy đoán ngày trễ.
    """
    status_date = status_date or date.today()
    try:
        end = datetime.strptime(str(end_s), DATE_FMT).date()
        actual = int(round(float(actual)))
    except Exception:
        return 0
    if actual >= 100:
        if not actual_finish_date:
            return 0
        try:
            finish = datetime.strptime(str(actual_finish_date), DATE_FMT).date()
        except Exception:
            return 0
        return max(0, (finish - end).days)
    return max(0, (status_date - end).days)


class Database:
    def __init__(self, path=DB_PATH):
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.create_tables()
        self.migrate()

    def create_tables(self):
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                start_date TEXT,
                end_date TEXT,
                manager TEXT,
                note TEXT DEFAULT '',
                source_mpp_path TEXT DEFAULT '',
                last_sync TEXT DEFAULT ''
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                wbs TEXT DEFAULT '',
                name TEXT NOT NULL,
                responsible TEXT DEFAULT '',
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                duration REAL DEFAULT 1,
                planned_progress INTEGER DEFAULT 0,
                actual_progress INTEGER DEFAULT 0,
                actual_override INTEGER DEFAULT NULL,
                actual_update_date TEXT DEFAULT '',
                actual_finish_date TEXT DEFAULT '',
                status TEXT DEFAULT 'Chưa bắt đầu',
                predecessor TEXT DEFAULT '',
                note TEXT DEFAULT '',
                source_type TEXT DEFAULT 'manual',
                source_uid INTEGER,
                source_task_id INTEGER,
                outline_level INTEGER DEFAULT 1,
                is_summary INTEGER DEFAULT 0,
                is_milestone INTEGER DEFAULT 0,
                critical INTEGER DEFAULT 0,
                total_slack REAL DEFAULT 0,
                resource_names TEXT DEFAULT '',
                baseline_start TEXT DEFAULT '',
                baseline_finish TEXT DEFAULT '',
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
        """)
        self.conn.commit()

    def _columns(self, table):
        return {r["name"] for r in self.conn.execute(f"PRAGMA table_info({table})")}

    def migrate(self):
        project_add = {
            "source_mpp_path": "TEXT DEFAULT ''",
            "last_sync": "TEXT DEFAULT ''",
        }
        task_add = {
            "source_type": "TEXT DEFAULT 'manual'",
            "source_uid": "INTEGER",
            "source_task_id": "INTEGER",
            "outline_level": "INTEGER DEFAULT 1",
            "is_summary": "INTEGER DEFAULT 0",
            "is_milestone": "INTEGER DEFAULT 0",
            "critical": "INTEGER DEFAULT 0",
            "total_slack": "REAL DEFAULT 0",
            "resource_names": "TEXT DEFAULT ''",
            "baseline_start": "TEXT DEFAULT ''",
            "baseline_finish": "TEXT DEFAULT ''",
            "actual_override": "INTEGER DEFAULT NULL",
            "actual_update_date": "TEXT DEFAULT ''",
            "actual_finish_date": "TEXT DEFAULT ''",
        }
        cols = self._columns("projects")
        for name, decl in project_add.items():
            if name not in cols:
                self.conn.execute(f"ALTER TABLE projects ADD COLUMN {name} {decl}")
        cols = self._columns("tasks")
        for name, decl in task_add.items():
            if name not in cols:
                self.conn.execute(f"ALTER TABLE tasks ADD COLUMN {name} {decl}")
        self.conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_tasks_mpp_uid "
            "ON tasks(project_id, source_type, source_uid) WHERE source_uid IS NOT NULL"
        )
        self.conn.commit()

    def projects(self):
        return self.conn.execute("SELECT * FROM projects ORDER BY id DESC").fetchall()

    def project(self, project_id):
        return self.conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()

    def add_project(self, code, name, start_date, end_date, manager, note="", source_mpp_path=""):
        cur = self.conn.execute(
            "INSERT INTO projects(code,name,start_date,end_date,manager,note,source_mpp_path) VALUES(?,?,?,?,?,?,?)",
            (code, name, start_date, end_date, manager, note, source_mpp_path)
        )
        self.conn.commit()
        return cur.lastrowid

    def update_project(self, project_id, code, name, start_date, end_date, manager, note):
        self.conn.execute(
            "UPDATE projects SET code=?,name=?,start_date=?,end_date=?,manager=?,note=? WHERE id=?",
            (code, name, start_date, end_date, manager, note, project_id)
        )
        self.conn.commit()

    def set_mpp_source(self, project_id, path, start_date="", end_date="", manager=""):
        values = [str(path), datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        sql = "UPDATE projects SET source_mpp_path=?, last_sync=?"
        if start_date:
            sql += ", start_date=?"; values.append(start_date)
        if end_date:
            sql += ", end_date=?"; values.append(end_date)
        if manager:
            sql += ", manager=?"; values.append(manager)
        sql += " WHERE id=?"; values.append(project_id)
        self.conn.execute(sql, values)
        self.conn.commit()

    def delete_project(self, project_id):
        self.conn.execute("DELETE FROM projects WHERE id=?", (project_id,))
        self.conn.commit()

    def tasks(self, project_id, keyword="", status="Tất cả"):
        sql = "SELECT * FROM tasks WHERE project_id=?"
        params = [project_id]
        if keyword:
            sql += " AND (name LIKE ? OR wbs LIKE ? OR responsible LIKE ? OR resource_names LIKE ?)"
            k = f"%{keyword}%"; params += [k, k, k, k]
        if status != "Tất cả":
            sql += " AND status=?"; params.append(status)
        # Microsoft Project ID is the most faithful display order after import.
        sql += " ORDER BY CASE WHEN source_type='mpp' THEN 0 ELSE 1 END, source_task_id, start_date, wbs, id"
        return self.conn.execute(sql, params).fetchall()

    def add_task(self, data):
        actual = max(0, min(100, int(data.get("actual_progress", 0) or 0)))
        if actual > 0 and not data.get("actual_update_date"):
            data["actual_update_date"] = date.today().isoformat()
        if actual >= 100 and not data.get("actual_finish_date"):
            data["actual_finish_date"] = date.today().isoformat()
        cols = [
            "project_id", "wbs", "name", "responsible", "start_date", "end_date", "duration",
            "planned_progress", "actual_progress", "actual_update_date", "actual_finish_date", "status", "predecessor", "note", "source_type",
            "source_uid", "source_task_id", "outline_level", "is_summary", "is_milestone", "critical",
            "total_slack", "resource_names", "baseline_start", "baseline_finish"
        ]
        vals = [data.get(c) for c in cols]
        self.conn.execute(
            f"INSERT INTO tasks({','.join(cols)}) VALUES({','.join(['?']*len(cols))})", vals
        )
        self.conn.commit()

    def update_task(self, task_id, data):
        old = self.get_task(task_id)
        actual = max(0, min(100, int(data.get("actual_progress", 0) or 0)))
        update_date = date.today().isoformat() if old and actual != int(old["actual_progress"] or 0) else (old["actual_update_date"] or "" if old else "")
        finish_date = old["actual_finish_date"] or "" if old else ""
        if actual >= 100 and not finish_date:
            finish_date = update_date or date.today().isoformat()
        elif actual < 100:
            finish_date = ""
        data["actual_update_date"] = update_date
        data["actual_finish_date"] = finish_date
        data["status"] = calc_progress_status(data["start_date"], data["end_date"], data["planned_progress"], actual)
        cols = ["wbs", "name", "responsible", "start_date", "end_date", "duration", "planned_progress",
                "actual_progress", "actual_update_date", "actual_finish_date", "status", "predecessor", "note"]
        self.conn.execute(
            f"UPDATE tasks SET {','.join(f'{c}=?' for c in cols)} WHERE id=?",
            [data[c] for c in cols] + [task_id]
        )
        self.conn.commit()

    def set_actual_override(self, task_id, actual, status_date=None):
        actual = max(0, min(100, int(actual)))
        status_date = status_date or date.today()
        task = self.get_task(task_id)
        if task is None:
            return "Chưa xác định", 0
        finish_date = task["actual_finish_date"] or ""
        if actual >= 100 and not finish_date:
            finish_date = status_date.isoformat()
        elif actual < 100:
            finish_date = ""
        status = calc_progress_status(task["start_date"], task["end_date"], task["planned_progress"], actual, status_date)
        self.conn.execute(
            "UPDATE tasks SET actual_progress=?, actual_override=?, actual_update_date=?, actual_finish_date=?, status=? WHERE id=?",
            (actual, actual, status_date.isoformat(), finish_date, status, task_id)
        )
        self.conn.commit()
        delay = calculate_delay_days(task["end_date"], actual, finish_date, status_date)
        return status, delay

    def clear_actual_override(self, task_id, actual_from_mpp, status):
        actual_from_mpp = max(0, min(100, int(actual_from_mpp)))
        self.conn.execute(
            "UPDATE tasks SET actual_progress=?, actual_override=NULL, actual_update_date='', actual_finish_date='', status=? WHERE id=?",
            (actual_from_mpp, status, task_id)
        )
        self.conn.commit()

    def delete_task(self, task_id):
        self.conn.execute("DELETE FROM tasks WHERE id=?", (task_id,)); self.conn.commit()

    def get_task(self, task_id):
        return self.conn.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()

    def sync_mpp_tasks(self, project_id, tasks):
        incoming_uids = set()
        with self.conn:
            for t in tasks:
                uid = int(t["source_uid"])
                incoming_uids.add(uid)
                payload = {
                    "project_id": project_id,
                    "wbs": t.get("wbs", ""),
                    "name": t.get("name", ""),
                    "responsible": "",
                    "start_date": t.get("start_date", ""),
                    "end_date": t.get("end_date", ""),
                    "duration": t.get("duration", 0),
                    "planned_progress": t.get("planned_progress", 0),
                    "actual_progress": t.get("actual_progress", 0),
                    "status": t.get("status", "Chưa xác định"),
                    "predecessor": t.get("predecessor", ""),
                    "note": t.get("note", ""),
                    "source_type": "mpp",
                    "source_uid": uid,
                    "source_task_id": t.get("task_id", 0),
                    "outline_level": t.get("outline_level", 1),
                    "is_summary": t.get("is_summary", 0),
                    "is_milestone": t.get("is_milestone", 0),
                    "critical": t.get("critical", 0),
                    "total_slack": t.get("total_slack", 0),
                    "resource_names": t.get("resource_names", ""),
                    "baseline_start": t.get("baseline_start", ""),
                    "baseline_finish": t.get("baseline_finish", ""),
                }
                row = self.conn.execute(
                    "SELECT id, actual_override FROM tasks WHERE project_id=? AND source_type='mpp' AND source_uid=?",
                    (project_id, uid)
                ).fetchone()
                cols = [k for k in payload if k != "project_id"]
                if row:
                    # TT% nhập trực tiếp trong app là override cục bộ.
                    # Đồng bộ MPP cập nhật các trường Project nhưng không ghi đè TT% override.
                    if row["actual_override"] is not None:
                        payload["actual_progress"] = int(row["actual_override"])
                        payload["status"] = calc_progress_status(
                            payload["start_date"], payload["end_date"],
                            payload["planned_progress"], payload["actual_progress"]
                        )
                    self.conn.execute(
                        f"UPDATE tasks SET {','.join(f'{c}=?' for c in cols)} WHERE id=?",
                        [payload[c] for c in cols] + [row["id"]]
                    )
                else:
                    all_cols = list(payload)
                    self.conn.execute(
                        f"INSERT INTO tasks({','.join(all_cols)}) VALUES({','.join(['?']*len(all_cols))})",
                        [payload[c] for c in all_cols]
                    )

            existing = self.conn.execute(
                "SELECT id, source_uid FROM tasks WHERE project_id=? AND source_type='mpp'", (project_id,)
            ).fetchall()
            for row in existing:
                if row["source_uid"] not in incoming_uids:
                    self.conn.execute("DELETE FROM tasks WHERE id=?", (row["id"],))


class ProjectDialog(QDialog):
    def __init__(self, parent=None, project=None):
        super().__init__(parent)
        self.setWindowTitle("Thông tin dự án"); self.setMinimumWidth(440)
        form = QFormLayout(self)
        self.code = QLineEdit(); self.name = QLineEdit()
        self.start = QDateEdit(calendarPopup=True); self.start.setDisplayFormat("dd/MM/yyyy"); self.start.setDate(QDate.currentDate())
        self.end = QDateEdit(calendarPopup=True); self.end.setDisplayFormat("dd/MM/yyyy"); self.end.setDate(QDate.currentDate().addMonths(12))
        self.manager = QLineEdit(); self.note = QTextEdit(); self.note.setFixedHeight(70)
        for label, widget in [("Mã dự án *", self.code), ("Tên dự án *", self.name), ("Ngày bắt đầu", self.start),
                              ("Ngày kết thúc", self.end), ("Chỉ huy/PM", self.manager), ("Ghi chú", self.note)]:
            form.addRow(label, widget)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_accept); buttons.rejected.connect(self.reject); form.addRow(buttons)
        if project:
            self.code.setText(project["code"]); self.name.setText(project["name"])
            if project["start_date"]: self.start.setDate(QDate.fromString(project["start_date"], "yyyy-MM-dd"))
            if project["end_date"]: self.end.setDate(QDate.fromString(project["end_date"], "yyyy-MM-dd"))
            self.manager.setText(project["manager"] or ""); self.note.setPlainText(project["note"] or "")

    def validate_accept(self):
        if not self.code.text().strip() or not self.name.text().strip():
            QMessageBox.warning(self, "Thiếu dữ liệu", "Vui lòng nhập mã và tên dự án."); return
        if self.end.date() < self.start.date():
            QMessageBox.warning(self, "Sai ngày", "Ngày kết thúc phải sau ngày bắt đầu."); return
        self.accept()

    def data(self):
        return dict(code=self.code.text().strip(), name=self.name.text().strip(),
                    start_date=self.start.date().toString("yyyy-MM-dd"), end_date=self.end.date().toString("yyyy-MM-dd"),
                    manager=self.manager.text().strip(), note=self.note.toPlainText().strip())


class TaskDialog(QDialog):
    def __init__(self, parent=None, task=None):
        super().__init__(parent)
        self.setWindowTitle("Công việc tiến độ"); self.setMinimumWidth(500)
        form = QFormLayout(self)
        self.wbs = QLineEdit(); self.name = QLineEdit(); self.responsible = QLineEdit()
        self.start = QDateEdit(calendarPopup=True); self.start.setDisplayFormat("dd/MM/yyyy"); self.start.setDate(QDate.currentDate())
        self.end = QDateEdit(calendarPopup=True); self.end.setDisplayFormat("dd/MM/yyyy"); self.end.setDate(QDate.currentDate().addDays(7))
        self.planned = QSpinBox(); self.planned.setRange(0, 100); self.planned.setSuffix(" %")
        self.actual = QSpinBox(); self.actual.setRange(0, 100); self.actual.setSuffix(" %")
        self.predecessor = QLineEdit(); self.note = QTextEdit(); self.note.setFixedHeight(70)
        rows = [("WBS", self.wbs), ("Tên công việc *", self.name), ("Phụ trách", self.responsible), ("Bắt đầu", self.start),
                ("Kết thúc", self.end), ("KH hoàn thành", self.planned), ("TT hoàn thành", self.actual),
                ("Công việc trước", self.predecessor), ("Ghi chú", self.note)]
        for x in rows: form.addRow(*x)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_accept); buttons.rejected.connect(self.reject); form.addRow(buttons)
        if task:
            self.wbs.setText(task["wbs"] or ""); self.name.setText(task["name"]); self.responsible.setText(task["responsible"] or "")
            self.start.setDate(QDate.fromString(task["start_date"], "yyyy-MM-dd")); self.end.setDate(QDate.fromString(task["end_date"], "yyyy-MM-dd"))
            self.planned.setValue(task["planned_progress"] or 0); self.actual.setValue(task["actual_progress"] or 0)
            self.predecessor.setText(task["predecessor"] or ""); self.note.setPlainText(task["note"] or "")

    def validate_accept(self):
        if not self.name.text().strip(): QMessageBox.warning(self, "Thiếu dữ liệu", "Vui lòng nhập tên công việc."); return
        if self.end.date() < self.start.date(): QMessageBox.warning(self, "Sai ngày", "Ngày kết thúc phải sau ngày bắt đầu."); return
        self.accept()

    @staticmethod
    def calc_status(start_s, end_s, planned, actual):
        return calc_progress_status(start_s, end_s, planned, actual)

    def data(self):
        start_s = self.start.date().toString("yyyy-MM-dd"); end_s = self.end.date().toString("yyyy-MM-dd")
        duration = self.start.date().daysTo(self.end.date()) + 1; planned = self.planned.value(); actual = self.actual.value()
        return dict(wbs=self.wbs.text().strip(), name=self.name.text().strip(), responsible=self.responsible.text().strip(),
                    start_date=start_s, end_date=end_s, duration=max(1, duration), planned_progress=planned, actual_progress=actual,
                    status=self.calc_status(start_s, end_s, planned, actual), predecessor=self.predecessor.text().strip(),
                    note=self.note.toPlainText().strip())


class StatCard(QFrame):
    def __init__(self, title):
        super().__init__(); self.setObjectName("card"); layout = QVBoxLayout(self)
        self.title = QLabel(title); self.title.setObjectName("cardTitle")
        self.value = QLabel("0"); self.value.setObjectName("cardValue")
        layout.addWidget(self.title); layout.addWidget(self.value)


class GanttView(QGraphicsView):
    def __init__(self):
        super().__init__()
        self.setScene(QGraphicsScene(self))
        self.setMinimumHeight(290)
        # Cache graphics theo DB task id để khi sửa TT% chỉ cập nhật đúng 1 task,
        # không phải dựng lại toàn bộ Gantt (rất chậm với vài trăm task).
        self._task_graphics = {}

    @staticmethod
    def _bar_color(task, status=None):
        if task["is_summary"]:
            return QColor("#334155")
        if task["critical"]:
            return QColor("#b91c1c")
        status = status if status is not None else task["status"]
        return {
            "Hoàn thành": QColor("#16a34a"),
            "Nhanh tiến độ": QColor("#16a34a"),
            "Đúng tiến độ": QColor("#2563eb"),
            "Đang thực hiện": QColor("#2563eb"),
            "Chậm tiến độ": QColor("#dc2626"),
            "Chưa bắt đầu": QColor("#94a3b8"),
        }.get(status, QColor("#64748b"))

    def draw_gantt(self, tasks):
        scene = self.scene()
        scene.clear()
        self._task_graphics = {}
        tasks = [t for t in tasks if t["start_date"] and t["end_date"]]
        if not tasks:
            scene.addText("Chưa có công việc để hiển thị Gantt.")
            return

        starts = [datetime.strptime(t["start_date"], DATE_FMT).date() for t in tasks]
        ends = [datetime.strptime(t["end_date"], DATE_FMT).date() for t in tasks]
        min_d, max_d = min(starts), max(ends)
        total_days = max(1, (max_d - min_d).days + 1)
        left, top, row_h = 300, 38, 28
        day_w = 18 if total_days <= 60 else (9 if total_days <= 150 else (5 if total_days <= 365 else 3))
        chart_w = total_days * day_w
        font = QFont(); font.setPointSize(8)

        for i in range(total_days):
            d = min_d + timedelta(days=i)
            x = left + i * day_w
            if d.weekday() == 0 or i == 0:
                scene.addLine(x, top - 6, x, top + len(tasks) * row_h, QPen(QColor("#d5dbe5")))
                txt = scene.addText(d.strftime("%d/%m"), font)
                txt.setPos(x + 2, 2)

        today = date.today()
        if min_d <= today <= max_d:
            x = left + (today - min_d).days * day_w
            scene.addLine(x, 0, x, top + len(tasks) * row_h, QPen(QColor("#ef4444"), 2))

        for row, t in enumerate(tasks):
            y = top + row * row_h
            indent = "   " * max(0, int(t["outline_level"] or 1) - 1)
            label = f"{t['wbs']}  {indent}{t['name']}".strip()
            txt = scene.addText(label[:62], font)
            txt.setPos(4, y + 2)

            s = datetime.strptime(t["start_date"], DATE_FMT).date()
            e = datetime.strptime(t["end_date"], DATE_FMT).date()
            x = left + (s - min_d).days * day_w
            w = max(day_w, ((e - s).days + 1) * day_w)

            if t["is_milestone"]:
                cx, cy, size = x, y + 13, 7
                poly = QPolygonF([
                    QPointF(cx, cy-size), QPointF(cx+size, cy),
                    QPointF(cx, cy+size), QPointF(cx-size, cy)
                ])
                item = QGraphicsPolygonItem(poly)
                item.setBrush(QBrush(QColor("#7c3aed")))
                item.setPen(QPen(Qt.NoPen))
                scene.addItem(item)
                self._task_graphics[int(t["id"])] = {"milestone": item}
                continue

            bar = QGraphicsRectItem(QRectF(x, y + 5, w, 16))
            bar.setBrush(QBrush(self._bar_color(t)))
            bar.setPen(QPen(Qt.NoPen))
            bar.setToolTip(
                f"{t['name']}\n{t['start_date']} → {t['end_date']}\n"
                f"TT: {t['actual_progress']}% | Trễ: {calculate_delay_days(t['end_date'], t['actual_progress'], t['actual_finish_date'] or '')} ngày\n"
                f"Critical: {'Có' if t['critical'] else 'Không'}"
            )
            scene.addItem(bar)

            pbar = None
            if not t["is_summary"]:
                pbar = QGraphicsRectItem(QRectF(x, y + 15, 0, 5))
                pbar.setBrush(QBrush(QColor("#0f172a")))
                pbar.setPen(QPen(Qt.NoPen))
                scene.addItem(pbar)
                progress = max(0, min(100, int(t["actual_progress"] or 0))) / 100.0
                pbar.setRect(QRectF(x, y + 15, w * progress, 5))
                pbar.setVisible(0 < progress < 1)

            self._task_graphics[int(t["id"])] = {
                "bar": bar,
                "progress": pbar,
                "x": x,
                "y": y,
                "w": w,
                "name": t["name"],
                "start": t["start_date"],
                "end": t["end_date"],
                "critical": bool(t["critical"]),
                "summary": bool(t["is_summary"]),
            }

        scene.setSceneRect(0, 0, left + chart_w + 60, top + len(tasks) * row_h + 30)

    def update_task_progress(self, task):
        """Cập nhật 1 task trên Gantt mà không redraw toàn bộ scene."""
        if task is None:
            return
        try:
            task_id = int(task["id"])
        except Exception:
            return
        g = self._task_graphics.get(task_id)
        if not g or "bar" not in g:
            return

        actual = max(0, min(100, int(task["actual_progress"] or 0)))
        status = task["status"] or ""
        bar = g["bar"]
        if not g.get("summary") and not g.get("critical"):
            bar.setBrush(QBrush(self._bar_color(task, status)))
        bar.setToolTip(
            f"{g['name']}\n{g['start']} → {g['end']}\n"
            f"TT: {actual}% | Trễ: {calculate_delay_days(task['end_date'], actual, task['actual_finish_date'] or '')} ngày\n"
            f"Critical: {'Có' if g.get('critical') else 'Không'}"
        )

        pbar = g.get("progress")
        if pbar is not None:
            progress = actual / 100.0
            pbar.setRect(QRectF(g["x"], g["y"] + 15, g["w"] * progress, 5))
            pbar.setVisible(0 < progress < 1)
        self.viewport().update()


class MainWindow(QMainWindow):
    STATUSES = ["Tất cả", "Chưa bắt đầu", "Đúng tiến độ", "Nhanh tiến độ", "Chậm tiến độ", "Hoàn thành", "Đang thực hiện", "Chưa xác định"]

    def __init__(self):
        super().__init__(); self.db = Database(); self.current_project_id = None
        self.setWindowTitle("QLDA Xây dựng V6.0 AI - Desktop • Tiến độ • Hồ sơ • Bản vẽ • Báo cáo • Văn bản • AI"); self.resize(1650, 930)
        self.build_ui(); self.apply_style(); self.load_projects()

    def build_ui(self):
        root = QWidget(); self.setCentralWidget(root)
        root_layout = QVBoxLayout(root); root_layout.setContentsMargins(8,8,8,8); root_layout.setSpacing(6)
        self.main_tabs = QTabWidget(); self.main_tabs.setDocumentMode(True)
        schedule_root = QWidget(); main = QVBoxLayout(schedule_root); main.setContentsMargins(12,10,12,10); main.setSpacing(10)
        self.main_tabs.addTab(schedule_root, "📅 Quản lý tiến độ")
        root_layout.addWidget(self.main_tabs, 1)

        header = QHBoxLayout(); title_box = QVBoxLayout()
        title = QLabel("QUẢN LÝ TIẾN ĐỘ DỰ ÁN"); title.setObjectName("pageTitle")
        subtitle = QLabel("QLDA Xây dựng • Microsoft Project MPP • WBS • Baseline • Critical Path • Gantt"); subtitle.setObjectName("subtitle")
        self.source_label = QLabel("Nguồn MPP: chưa liên kết"); self.source_label.setObjectName("sourceLabel"); self.source_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        title_box.addWidget(title); title_box.addWidget(subtitle); title_box.addWidget(self.source_label)
        header.addLayout(title_box); header.addStretch()
        self.project_combo = QComboBox(); self.project_combo.setMinimumWidth(360); self.project_combo.currentIndexChanged.connect(self.on_project_changed)
        for text, slot, obj in [("+ Dự án", self.add_project, ""), ("Sửa dự án", self.edit_project, ""), ("Xóa dự án", self.delete_project, "dangerBtn")]:
            b=QPushButton(text); b.clicked.connect(slot); b.setObjectName(obj); header.addWidget(b) if False else None
            if text == "+ Dự án": self.btn_project=b
            elif text == "Sửa dự án": self.btn_edit_project=b
            else: self.btn_del_project=b
        header.addWidget(QLabel("Dự án:")); header.addWidget(self.project_combo); header.addWidget(self.btn_project); header.addWidget(self.btn_edit_project); header.addWidget(self.btn_del_project)
        main.addLayout(header)

        cards = QGridLayout(); self.card_total=StatCard("Tổng công việc"); self.card_delay=StatCard("Chậm tiến độ"); self.card_critical=StatCard("Critical"); self.card_done=StatCard("Hoàn thành"); self.card_avg=StatCard("Tiến độ TB")
        for i,c in enumerate([self.card_total,self.card_delay,self.card_critical,self.card_done,self.card_avg]): cards.addWidget(c,0,i)
        main.addLayout(cards)

        mpp_controls = QHBoxLayout()
        self.btn_mpp = QPushButton("📂 Mở / Liên kết MPP"); self.btn_mpp.setObjectName("mppBtn"); self.btn_mpp.clicked.connect(self.link_mpp)
        self.btn_sync = QPushButton("⟳ Đồng bộ MPP"); self.btn_sync.setObjectName("mppBtn"); self.btn_sync.clicked.connect(self.sync_mpp)
        self.btn_open_project = QPushButton("Mở bằng Microsoft Project"); self.btn_open_project.clicked.connect(self.open_in_ms_project)
        mpp_controls.addWidget(self.btn_mpp); mpp_controls.addWidget(self.btn_sync); mpp_controls.addWidget(self.btn_open_project); mpp_controls.addStretch()
        main.addLayout(mpp_controls)

        controls = QHBoxLayout()
        self.btn_add_task=QPushButton("+ Thêm thủ công"); self.btn_add_task.clicked.connect(self.add_task)
        self.btn_edit_task=QPushButton("Sửa"); self.btn_edit_task.clicked.connect(self.edit_task)
        self.btn_delete_task=QPushButton("Xóa"); self.btn_delete_task.setObjectName("dangerBtn"); self.btn_delete_task.clicked.connect(self.delete_task)
        self.btn_reset_tt=QPushButton("TT từ MPP"); self.btn_reset_tt.setToolTip("Bỏ TT% nhập tay và lấy lại % Complete từ Microsoft Project ở lần đồng bộ tiếp theo")
        self.btn_reset_tt.clicked.connect(self.reset_actual_override)
        self.search=QLineEdit(); self.search.setPlaceholderText("Tìm WBS / công việc / resource..."); self.search.textChanged.connect(self.refresh_tasks)
        self.status_filter=QComboBox(); self.status_filter.addItems(self.STATUSES); self.status_filter.currentTextChanged.connect(self.refresh_tasks)
        btn_import=QPushButton("Nhập Excel"); btn_import.clicked.connect(self.import_excel); btn_export=QPushButton("Xuất Excel"); btn_export.clicked.connect(self.export_excel)
        for w in [self.btn_add_task,self.btn_edit_task,self.btn_delete_task,self.btn_reset_tt]: controls.addWidget(w)
        controls.addSpacing(8); controls.addWidget(self.search,1); controls.addWidget(self.status_filter); controls.addWidget(btn_import); controls.addWidget(btn_export); main.addLayout(controls)
        tt_hint = QLabel("Mẹo: Double-click ô TT % → nhập 0–100. TT=100% là Hoàn thành; nếu quá ngày Kết thúc, cột Ngày trễ tự tính số ngày chậm. TT nhập tay được giữ khi đồng bộ MPP.")
        tt_hint.setObjectName("ttHint"); main.addWidget(tt_hint)

        splitter=QSplitter(Qt.Vertical)
        self.table=QTableWidget(0,18); self.table.setHorizontalHeaderLabels([
            "DB ID","Project ID","WBS","Công việc","Bắt đầu","Kết thúc","Duration","KH %","TT %","Nhanh / Chậm","Trạng thái","Ngày trễ",
            "Predecessor","Resources","Baseline Start","Baseline Finish","Slack (ngày)","Critical"
        ])
        self.table.setColumnHidden(0, True); self.table.setSelectionBehavior(QAbstractItemView.SelectRows); self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.verticalHeader().setVisible(False); self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents); self.table.horizontalHeader().setSectionResizeMode(3,QHeaderView.Stretch)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        self.table.itemChanged.connect(self.on_table_item_changed)
        splitter.addWidget(self.table)
        self.gantt=GanttView(); splitter.addWidget(self.gantt); splitter.setSizes([470,310]); main.addWidget(splitter,1)

        # Sheet cấp 1 thứ hai: Quản lý hồ sơ. Bên trong có các sheet nhỏ NCR/RFA/RFI/VO.
        self.documents_page = DocumentManagerPage(self.db, self)
        self.documents_page.projectChanged.connect(self.on_document_project_changed)
        self.main_tabs.addTab(self.documents_page, "📁 Quản lý hồ sơ")

        # Sheet cấp 1 thứ ba: Quản lý bản vẽ.
        self.drawings_page = DrawingManagerPage(self.db, self)
        self.drawings_page.projectChanged.connect(self.on_drawing_project_changed)
        self.main_tabs.addTab(self.drawings_page, "📐 Quản lý bản vẽ")

        # Sheet cấp 1 thứ tư: Dashboard báo cáo trực quan.
        self.report_page = ReportDashboardPage(self.db, self)
        self.report_tab_index = self.main_tabs.addTab(self.report_page, "📊 Báo cáo trực quan")

        # Sheet cấp 1 thứ năm: Văn bản QLDA xây dựng, cập nhật online.
        self.legal_page = LegalDocumentsPage(DB_PATH, self)
        self.legal_tab_index = self.main_tabs.addTab(self.legal_page, "📚 Văn bản QLDA XD")

        # Sheet cấp 1 thứ sáu: Trợ lý AI.
        self.ai_page = AIAssistantPage(DB_PATH, self)
        self.ai_tab_index = self.main_tabs.addTab(self.ai_page, "🤖 Trợ lý AI")

        # Sheet cấp 1 thứ bảy: Cài đặt tập trung.
        self.settings_page = SettingsPage(DB_PATH, self)
        self.settings_tab_index = self.main_tabs.addTab(self.settings_page, "⚙ Cài đặt")
        self.settings_page.settingsChanged.connect(self.legal_page.reload_settings_status)
        self.settings_page.settingsChanged.connect(self.ai_page.refresh_settings_status)
        self.settings_page.settingsChanged.connect(self.refresh_access_role)

        self.main_tabs.currentChanged.connect(self.on_main_tab_changed)
        self.statusBar().showMessage("Sẵn sàng")
        self.refresh_access_role()

    def refresh_access_role(self):
        self.access = desktop_access_state()
        role = self.access.role
        can_update = role in {"update", "admin"}
        is_admin = role == "admin"
        # Project master data is Admin-only; progress content can be updated by Cập nhật/Admin.
        for b in (self.btn_project, self.btn_edit_project, self.btn_del_project):
            b.setEnabled(is_admin)
        for b in (self.btn_mpp, self.btn_sync, self.btn_add_task, self.btn_edit_task, self.btn_reset_tt):
            b.setEnabled(can_update)
        self.btn_delete_task.setEnabled(is_admin)
        self.table.setEditTriggers(
            (QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed) if can_update else QAbstractItemView.NoEditTriggers
        )
        if hasattr(self, "documents_page"):
            self.documents_page.set_access_role(role)
        if hasattr(self, "drawings_page"):
            self.drawings_page.set_access_role(role)
        if self.access.drive_enabled:
            self.statusBar().showMessage(f"Google Drive • {self.access.label}")
        else:
            self.statusBar().showMessage("Local mode • Admin")

    def _require_update(self) -> bool:
        access = getattr(self, "access", desktop_access_state())
        if access.role not in {"update", "admin"}:
            QMessageBox.warning(self, "Chỉ đọc", "Tài khoản hiện tại chỉ có quyền đọc. Không thể cập nhật dữ liệu.")
            return False
        return True

    def _require_admin(self) -> bool:
        access = getattr(self, "access", desktop_access_state())
        if access.role != "admin":
            QMessageBox.warning(self, "Cần quyền Admin", "Chức năng này yêu cầu quyền Admin.")
            return False
        return True

    def apply_style(self):
        self.setStyleSheet("""
            QMainWindow, QWidget { background:#f4f7fb; color:#172033; font-size:13px; }
            QLabel#pageTitle { font-size:23px; font-weight:800; color:#0f172a; }
            QLabel#subtitle { color:#64748b; } QLabel#sourceLabel { color:#0f766e; font-size:12px; }
            QLabel#ttHint { color:#92400e; background:#fffbeb; border:1px solid #fde68a; border-radius:5px; padding:6px 9px; }
            QFrame#card, QFrame#docCard { background:white; border:1px solid #e2e8f0; border-radius:10px; }
            QLabel#cardTitle, QLabel#docCardTitle { color:#64748b; font-size:12px; }
            QLabel#cardValue, QLabel#docCardValue { color:#0f172a; font-size:24px; font-weight:800; }
            QLabel#docHint { color:#1e3a5f; background:#eff6ff; border:1px solid #bfdbfe; border-radius:5px; padding:7px 9px; }
            QLabel#drawingHint { color:#14532d; background:#f0fdf4; border:1px solid #bbf7d0; border-radius:5px; padding:7px 9px; }
            QLabel#reportPageTitle { font-size:22px; font-weight:800; color:#0f172a; }
            QLabel#reportProjectLabel { color:#64748b; font-size:12px; }
            QFrame#reportMetricCard, QFrame#reportChartPanel { background:white; border:1px solid #dbe3ef; border-radius:9px; }
            QLabel#reportMetricTitle { color:#64748b; font-size:12px; font-weight:600; }
            QLabel#reportMetricValue { color:#0f172a; font-size:24px; font-weight:800; }
            QLabel#reportMetricNote { color:#64748b; font-size:11px; }
            QLabel#reportChartTitle { color:#0f172a; font-size:15px; font-weight:800; }
            QLabel#reportChartSubtitle { color:#64748b; font-size:11px; }
            QFrame#legalCard { background:white; border:1px solid #dbe3ef; border-radius:9px; }
            QLabel#legalCardTitle { color:#64748b; font-size:12px; font-weight:600; }
            QLabel#legalCardValue { color:#0f172a; font-size:24px; font-weight:800; }
            QLabel#aiPageTitle { font-size:22px; font-weight:800; color:#0f172a; }
            QLabel#aiSubtitle { color:#64748b; }
            QLabel#aiHint { color:#312e81; background:#eef2ff; border:1px solid #c7d2fe; border-radius:5px; padding:7px 9px; }
            QFrame#aiConfig { background:white; border:1px solid #dbe3ef; border-radius:8px; }
            QTabWidget::pane { border:1px solid #dbe3ef; background:#f8fafc; border-radius:6px; }
            QTabBar::tab { background:#eaf1fb; color:#1e3a5f; padding:9px 18px; margin-right:2px; font-weight:700; }
            QTabBar::tab:selected { background:#0f62fe; color:white; }
            QPushButton { background:#0f62fe; color:white; border:none; border-radius:6px; padding:8px 12px; font-weight:600; }
            QPushButton:hover { background:#0043ce; } QPushButton#dangerBtn { background:#dc2626; } QPushButton#dangerBtn:hover { background:#b91c1c; }
            QPushButton#mppBtn { background:#0f766e; } QPushButton#mppBtn:hover { background:#115e59; }
            QLineEdit,QComboBox,QDateEdit,QSpinBox,QTextEdit { background:white; border:1px solid #cbd5e1; border-radius:5px; padding:6px; }
            QTableWidget { background:white; alternate-background-color:#f8fafc; border:1px solid #dbe3ef; gridline-color:#e8edf5; }
            QHeaderView::section { background:#eaf1fb; color:#1e3a5f; padding:7px; border:none; border-right:1px solid #d7e0ec; font-weight:700; }
        """); self.table.setAlternatingRowColors(True)

    def load_projects(self, select_id=None):
        old = select_id or self.current_project_id; self.project_combo.blockSignals(True); self.project_combo.clear(); projects=self.db.projects()
        for p in projects: self.project_combo.addItem(f"{p['code']} - {p['name']}", p['id'])
        self.project_combo.blockSignals(False)
        if projects:
            idx=0
            if old:
                for i in range(self.project_combo.count()):
                    if self.project_combo.itemData(i)==old: idx=i; break
            self.project_combo.setCurrentIndex(idx); self.on_project_changed(idx)
        else:
            self.current_project_id=None; self.refresh_tasks(); self.update_source_label()
        if hasattr(self, "documents_page"):
            self.documents_page.reload_projects(self.current_project_id)
        if hasattr(self, "drawings_page"):
            self.drawings_page.reload_projects(self.current_project_id)
        if hasattr(self, "report_page"):
            self.report_page.set_project(self.current_project_id)

    def on_project_changed(self,index):
        self.current_project_id=self.project_combo.itemData(index) if index>=0 else None
        self.update_source_label(); self.refresh_tasks()
        if hasattr(self, "documents_page") and self.current_project_id:
            self.documents_page.set_project(self.current_project_id)
        if hasattr(self, "drawings_page") and self.current_project_id:
            self.drawings_page.set_project(self.current_project_id)
        if hasattr(self, "report_page"):
            self.report_page.set_project(self.current_project_id)
        if hasattr(self, "ai_page"):
            self.ai_page.set_project(self.current_project_id)

    def on_main_tab_changed(self, index):
        if index == self.settings_tab_index:
            self.settings_page.refresh_drive_status(silent=True)
        if hasattr(self, "report_page") and index == getattr(self, "report_tab_index", -1):
            self.report_page.set_project(self.current_project_id)
        if hasattr(self, "legal_page") and index == getattr(self, "legal_tab_index", -1):
            self.legal_page.reload_settings_status()
        if hasattr(self, "ai_page") and index == getattr(self, "ai_tab_index", -1):
            self.ai_page.refresh_settings_status()
            self.ai_page.set_project(self.current_project_id)
        if hasattr(self, "settings_page") and index == getattr(self, "settings_tab_index", -1):
            self.settings_page.load_values()

    def on_document_project_changed(self, project_id):
        """Giữ lựa chọn dự án đồng bộ giữa sheet Tiến độ và sheet Hồ sơ."""
        if not project_id or project_id == self.current_project_id:
            return
        for i in range(self.project_combo.count()):
            if self.project_combo.itemData(i) == project_id:
                self.project_combo.setCurrentIndex(i)
                return

    def on_drawing_project_changed(self, project_id):
        """Giữ lựa chọn dự án đồng bộ giữa Tiến độ, Hồ sơ và Bản vẽ."""
        if not project_id or project_id == self.current_project_id:
            return
        for i in range(self.project_combo.count()):
            if self.project_combo.itemData(i) == project_id:
                self.project_combo.setCurrentIndex(i)
                return

    def update_source_label(self):
        if not self.current_project_id:
            self.source_label.setText("Nguồn MPP: chưa chọn dự án"); return
        p=self.db.project(self.current_project_id); path=(p["source_mpp_path"] or "").strip(); sync=(p["last_sync"] or "").strip()
        if path: self.source_label.setText(f"Nguồn MPP: {path}   |   Đồng bộ gần nhất: {sync or 'chưa có'}")
        else: self.source_label.setText("Nguồn MPP: chưa liên kết")

    def add_project(self):
        if not self._require_admin(): return
        dlg=ProjectDialog(self)
        if dlg.exec()==QDialog.Accepted:
            try: pid=self.db.add_project(**dlg.data()); self.load_projects(pid)
            except sqlite3.IntegrityError: QMessageBox.warning(self,"Trùng mã","Mã dự án đã tồn tại.")

    def edit_project(self):
        if not self._require_admin(): return
        if not self.current_project_id: return
        p=self.db.project(self.current_project_id); dlg=ProjectDialog(self,p)
        if dlg.exec()==QDialog.Accepted:
            try: self.db.update_project(self.current_project_id,**dlg.data()); self.load_projects(self.current_project_id)
            except sqlite3.IntegrityError: QMessageBox.warning(self,"Trùng mã","Mã dự án đã tồn tại.")

    def delete_project(self):
        if self.current_project_id and QMessageBox.question(self,"Xác nhận","Xóa dự án và toàn bộ công việc, hồ sơ, bản vẽ?")==QMessageBox.Yes:
            self.db.delete_project(self.current_project_id); self.current_project_id=None; self.load_projects()

    @staticmethod
    def _code_from_filename(path):
        stem=Path(path).stem.upper(); asciiish=re.sub(r"[^A-Z0-9]+","-",stem); asciiish=asciiish.strip("-") or "MPP"
        return asciiish[:30]

    def _unique_project_code(self, base):
        code=base; n=2
        while self.db.conn.execute("SELECT 1 FROM projects WHERE code=?",(code,)).fetchone():
            suffix=f"-{n}"; code=(base[:30-len(suffix)] + suffix); n+=1
        return code

    def link_mpp(self):
        if not self._require_update(): return
        path,_=QFileDialog.getOpenFileName(self,"Chọn file Microsoft Project","","Microsoft Project (*.mpp *.mpt)")
        if not path: return
        self._read_and_sync_mpp(path, allow_create=True)

    def sync_mpp(self):
        if not self._require_update(): return
        if not self.current_project_id:
            QMessageBox.information(self,"Chưa có dự án","Hãy mở/liên kết một file MPP trước."); return
        p=self.db.project(self.current_project_id); path=(p["source_mpp_path"] or "").strip()
        if not path:
            self.link_mpp(); return
        if not Path(path).exists():
            QMessageBox.warning(self,"Không thấy file",f"Đường dẫn MPP cũ không còn tồn tại:\n{path}\n\nHãy chọn lại file."); self.link_mpp(); return
        self._read_and_sync_mpp(path, allow_create=False)

    def _read_and_sync_mpp(self,path,allow_create):
        QApplication.setOverrideCursor(Qt.WaitCursor); self.statusBar().showMessage("Đang đọc Microsoft Project...")
        try:
            data=MppComReader().read(path)
        except MppReadError as exc:
            QMessageBox.critical(self,"Không đọc được MPP",str(exc)); return
        except Exception as exc:
            QMessageBox.critical(self,"Lỗi MPP",f"Lỗi không xác định khi đọc MPP:\n{exc}"); return
        finally:
            QApplication.restoreOverrideCursor()
        if not data["tasks"]:
            QMessageBox.warning(self,"MPP không có công việc","Đã mở file nhưng không lấy được task. Hãy chạy test_project_com.py nếu cần chẩn đoán."); return
        if self.current_project_id:
            pid=self.current_project_id
            # If user selected another MPP while a project is active, ask before replacing source.
            p=self.db.project(pid); old=(p["source_mpp_path"] or "").strip()
            if allow_create and old and Path(old) != Path(path):
                ans=QMessageBox.question(self,"Đổi nguồn MPP","Dự án hiện tại đã liên kết MPP khác.\nDùng file mới làm nguồn cho dự án này?")
                if ans!=QMessageBox.Yes:
                    pid=None
            if pid is None:
                code=self._unique_project_code(self._code_from_filename(path)); pid=self.db.add_project(code,Path(path).stem,data["project_start"] or date.today().isoformat(),data["project_finish"] or date.today().isoformat(),data["manager"],"Tạo từ Microsoft Project",str(path))
        else:
            code=self._unique_project_code(self._code_from_filename(path)); pid=self.db.add_project(code,Path(path).stem,data["project_start"] or date.today().isoformat(),data["project_finish"] or date.today().isoformat(),data["manager"],"Tạo từ Microsoft Project",str(path))
        self.db.sync_mpp_tasks(pid,data["tasks"]); self.db.set_mpp_source(pid,path,data["project_start"],data["project_finish"],data["manager"])
        self.current_project_id=pid; self.load_projects(pid); self.statusBar().showMessage(f"Đã đồng bộ {len(data['tasks'])} task từ MPP")
        diag=data.get("diagnostics", {})
        extra=""
        if diag.get("uid_fallback"):
            extra=f"\n\nLưu ý: {diag['uid_fallback']} task phải dùng ID thay cho UniqueID."
        QMessageBox.information(self,"Đồng bộ hoàn tất",f"Đã đọc {len(data['tasks'])} công việc từ Microsoft Project.{extra}\n\nFile nguồn:\n{path}")

    def open_in_ms_project(self):
        if not self.current_project_id: return
        p=self.db.project(self.current_project_id); path=(p["source_mpp_path"] or "").strip()
        if not path: QMessageBox.information(self,"Chưa liên kết","Dự án chưa có file MPP nguồn."); return
        try:
            if os.name=="nt": os.startfile(path)
            else: QMessageBox.information(self,"Windows","Chức năng này dùng trên Windows có Microsoft Project.")
        except Exception as exc: QMessageBox.critical(self,"Không mở được",str(exc))

    def add_task(self):
        if not self._require_update(): return
        if not self.current_project_id: QMessageBox.information(self,"Chưa có dự án","Hãy tạo/chọn dự án trước."); return
        dlg=TaskDialog(self)
        if dlg.exec()==QDialog.Accepted:
            d=dlg.data(); d.update(project_id=self.current_project_id,source_type="manual",source_uid=None,source_task_id=None,outline_level=1,is_summary=0,is_milestone=0,critical=0,total_slack=0,resource_names="",baseline_start="",baseline_finish="")
            self.db.add_task(d); self.refresh_tasks()

    def selected_task_id(self):
        r=self.table.currentRow()
        try: return int(self.table.item(r,0).text()) if r>=0 else None
        except Exception: return None

    def edit_task(self):
        tid=self.selected_task_id()
        if not tid: return
        task=self.db.get_task(tid)
        if task["source_type"]=="mpp":
            QMessageBox.information(self,"Task từ Microsoft Project","Các trường WBS/ngày/resource của task MPP được quản lý trong Microsoft Project.\n\nRiêng cột TT % có thể double-click và nhập trực tiếp ngay trên bảng; app sẽ tự tính nhanh/chậm và giữ giá trị này khi đồng bộ MPP."); return
        dlg=TaskDialog(self,task)
        if dlg.exec()==QDialog.Accepted: self.db.update_task(tid,dlg.data()); self.refresh_tasks()

    def on_cell_double_clicked(self, row, column):
        # TT % (cột 8) được sửa trực tiếp. Các cột khác dùng hộp thoại cũ.
        if column == 8:
            return
        self.edit_task()

    def _update_progress_row_fast(self, row, planned, actual, status, delay_days=0, override=True):
        """Chỉ cập nhật các ô liên quan của 1 dòng, tránh reload 600+ task."""
        evaluation = progress_evaluation(planned, actual)
        self._loading_table = True
        self.table.blockSignals(True)
        try:
            tt_item = self.table.item(row, 8)
            eval_item = self.table.item(row, 9)
            status_item = self.table.item(row, 10)
            delay_item = self.table.item(row, 11)
            if tt_item is not None:
                tt_item.setText(str(actual))
                tt_item.setBackground(QColor("#fde68a" if override else "#fff7cc"))
            if eval_item is not None:
                eval_item.setText(evaluation)
                delta = progress_delta(planned, actual)
                eval_item.setBackground(QColor("#fee2e2" if delta < -1 else ("#dcfce7" if delta > 1 else "#e0f2fe")))
            if status_item is not None:
                status_item.setText(status)
                color = {
                    "Hoàn thành":"#dcfce7", "Đúng tiến độ":"#dbeafe", "Nhanh tiến độ":"#dcfce7",
                    "Đang thực hiện":"#dbeafe", "Chậm tiến độ":"#fee2e2", "Chưa bắt đầu":"#e2e8f0"
                }.get(status, "#fef3c7")
                status_item.setBackground(QColor(color))
            if delay_item is not None:
                delay_item.setText(str(int(delay_days)))
                delay_item.setBackground(QColor("#fee2e2" if delay_days > 0 else "#f8fafc"))
                delay_item.setToolTip("Số ngày vượt quá ngày Kết thúc. Khi TT=100%, số ngày trễ được khóa tại ngày đạt 100%.")
        finally:
            self.table.blockSignals(False)
            self._loading_table = False

    def on_table_item_changed(self, item):
        if item.column() != 8 or getattr(self, "_loading_table", False):
            return

        row = item.row()
        id_item = self.table.item(row, 0)
        kh_item = self.table.item(row, 7)
        if id_item is None or kh_item is None:
            return

        try:
            task_id = int(id_item.text())
        except Exception:
            return
        task = self.db.get_task(task_id)
        if task is None:
            return
        old_actual = int(task["actual_progress"] or 0)

        try:
            raw = item.text().strip().replace("%", "").replace(",", ".")
            actual = int(round(float(raw)))
        except Exception:
            QMessageBox.warning(self, "TT % không hợp lệ", "TT % phải là số từ 0 đến 100.")
            planned = int(float(kh_item.text() or 0))
            old_delay = calculate_delay_days(task["end_date"], old_actual, task["actual_finish_date"] or "")
            self._update_progress_row_fast(row, planned, old_actual, task["status"] or "", old_delay, task["actual_override"] is not None)
            return

        if actual < 0 or actual > 100:
            QMessageBox.warning(self, "TT % không hợp lệ", "TT % phải nằm trong khoảng 0–100.")
            planned = int(float(kh_item.text() or 0))
            old_delay = calculate_delay_days(task["end_date"], old_actual, task["actual_finish_date"] or "")
            self._update_progress_row_fast(row, planned, old_actual, task["status"] or "", old_delay, task["actual_override"] is not None)
            return

        planned = int(float(kh_item.text() or 0))

        # 1) Lưu DB ngay, đồng thời ghi ngày cập nhật/ngày đạt 100% và tính ngày trễ.
        status, delay_days = self.db.set_actual_override(task_id, actual, date.today())

        # 2) Chỉ cập nhật 1 dòng trên bảng, không gọi refresh_tasks().
        self._update_progress_row_fast(row, planned, actual, status, delay_days, True)

        # 3) Chỉ cập nhật đúng thanh Gantt của task đang sửa.
        updated_task = self.db.get_task(task_id)
        self.gantt.update_task_progress(updated_task)

        # 4) KPI chỉ cần query + tính tổng; không rebuild QTable/QGraphicsScene.
        tasks = self.db.tasks(
            self.current_project_id,
            self.search.text().strip(),
            self.status_filter.currentText()
        )
        self.update_cards(tasks)

        self.statusBar().showMessage(
            f"Đã cập nhật TT {actual}% | KH {planned}% | {progress_evaluation(planned, actual)} | Trễ {delay_days} ngày", 5000
        )

    def reset_actual_override(self):
        if not self._require_update(): return
        task_id = self.selected_task_id()
        if not task_id:
            return
        task = self.db.get_task(task_id)
        if task["source_type"] != "mpp":
            QMessageBox.information(self, "Task thủ công", "Task này không có dữ liệu % Complete từ MPP để khôi phục.")
            return
        # Bỏ override. Giá trị % Complete chuẩn sẽ được nạp lại khi Đồng bộ MPP.
        self.db.conn.execute("UPDATE tasks SET actual_override=NULL, actual_update_date='', actual_finish_date='' WHERE id=?", (task_id,))
        self.db.conn.commit()
        QMessageBox.information(self, "Đã bỏ TT nhập tay", "Đã bỏ override TT %. Bấm 'Đồng bộ MPP' để lấy lại % Complete từ Microsoft Project.")

    def delete_task(self):
        tid=self.selected_task_id()
        if not tid: return
        task=self.db.get_task(tid)
        if task["source_type"]=="mpp":
            QMessageBox.information(self,"Task từ MPP","Không xóa task MPP trong app. Hãy xóa trong Microsoft Project rồi Đồng bộ MPP."); return
        if QMessageBox.question(self,"Xác nhận","Xóa công việc đang chọn?")==QMessageBox.Yes: self.db.delete_task(tid); self.refresh_tasks()

    def refresh_tasks(self):
        if not self.current_project_id:
            self.table.setRowCount(0); self.gantt.draw_gantt([]); self.update_cards([]); return
        tasks=self.db.tasks(self.current_project_id,self.search.text().strip(),self.status_filter.currentText())
        self._loading_table = True
        self.table.blockSignals(True)
        self.table.setUpdatesEnabled(False)
        try:
            self.table.setRowCount(len(tasks))
            for r,t in enumerate(tasks):
                indent="   "*max(0,int(t["outline_level"] or 1)-1)
                name=("◆ " if t["is_milestone"] else "") + indent + t["name"]
                planned=int(t["planned_progress"] or 0); actual=int(t["actual_progress"] or 0)
                evaluation=progress_evaluation(planned, actual)
                status=calc_progress_status(t["start_date"], t["end_date"], planned, actual)
                delay_days=calculate_delay_days(t["end_date"], actual, t["actual_finish_date"] or "")
                # Giữ trạng thái DB đồng bộ với phép tính mới khi có khác biệt.
                if status != (t["status"] or ""):
                    self.db.conn.execute("UPDATE tasks SET status=? WHERE id=?", (status, t["id"]))
                vals=[t["id"],t["source_task_id"] or "",t["wbs"],name,t["start_date"],t["end_date"],t["duration"],planned,actual,evaluation,status,delay_days,t["predecessor"],t["resource_names"],t["baseline_start"],t["baseline_finish"],t["total_slack"],"YES" if t["critical"] else ""]
                for c,v in enumerate(vals):
                    item=QTableWidgetItem(str(v if v is not None else ""))
                    if c in (1,6,7,8,9,11,16,17): item.setTextAlignment(Qt.AlignCenter)
                    # Chỉ cột TT % được phép sửa trực tiếp.
                    if c == 8:
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                        item.setBackground(QColor("#fff7cc" if t["actual_override"] is None else "#fde68a"))
                        item.setToolTip("Double-click để nhập TT % (0–100). Màu vàng đậm = đang dùng TT nhập tay/override.")
                    else:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    if t["is_summary"] and c in (2,3):
                        font=item.font(); font.setBold(True); item.setFont(font)
                    self.table.setItem(r,c,item)
                # Màu đánh giá nhanh/chậm
                delta=progress_delta(planned,actual)
                if delta < -1:
                    self.table.item(r,9).setBackground(QColor("#fee2e2"))
                elif delta > 1:
                    self.table.item(r,9).setBackground(QColor("#dcfce7"))
                else:
                    self.table.item(r,9).setBackground(QColor("#e0f2fe"))
                color={"Hoàn thành":"#dcfce7","Đúng tiến độ":"#dbeafe","Nhanh tiến độ":"#dcfce7","Đang thực hiện":"#dbeafe","Chậm tiến độ":"#fee2e2","Chưa bắt đầu":"#e2e8f0"}.get(status,"#fef3c7")
                self.table.item(r,10).setBackground(QColor(color))
                self.table.item(r,11).setBackground(QColor("#fee2e2" if delay_days > 0 else "#f8fafc"))
                self.table.item(r,11).setToolTip("Số ngày vượt quá ngày Kết thúc. TT=100% khóa số ngày trễ tại ngày hoàn thành.")
                if t["critical"]: self.table.item(r,17).setBackground(QColor("#fecaca"))
            self.db.conn.commit()
        finally:
            self.table.setUpdatesEnabled(True)
            self.table.blockSignals(False)
            self._loading_table = False
        # Lấy lại rows sau khi status có thể đã được cập nhật để Gantt/card dùng trạng thái mới.
        tasks=self.db.tasks(self.current_project_id,self.search.text().strip(),self.status_filter.currentText())
        self.gantt.draw_gantt(tasks); self.update_cards(tasks); self.statusBar().showMessage(f"Đang hiển thị {len(tasks)} công việc")

    def update_cards(self,tasks):
        total=len(tasks); delay=sum(1 for t in tasks if t["status"]=="Chậm tiến độ"); critical=sum(1 for t in tasks if t["critical"]); done=sum(1 for t in tasks if t["status"]=="Hoàn thành"); avg=round(sum((t["actual_progress"] or 0) for t in tasks)/total,1) if total else 0
        self.card_total.value.setText(str(total)); self.card_delay.value.setText(str(delay)); self.card_critical.value.setText(str(critical)); self.card_done.value.setText(str(done)); self.card_avg.value.setText(f"{avg}%")

    def export_excel(self):
        if Workbook is None: QMessageBox.warning(self,"Thiếu thư viện","Cài openpyxl: pip install openpyxl"); return
        if not self.current_project_id: return
        path,_=QFileDialog.getSaveFileName(self,"Xuất tiến độ","Tien_do_du_an.xlsx","Excel (*.xlsx)")
        if not path: return
        tasks=self.db.tasks(self.current_project_id); wb=Workbook(); ws=wb.active; ws.title="TienDo"
        headers=["Project ID","UID","WBS","Công việc","Bắt đầu","Kết thúc","Duration","KH %","TT %","Nhanh/Chậm","Lệch TT-KH (%)","Trạng thái","Ngày trễ","Ngày cập nhật TT","Ngày đạt 100%","Predecessor","Resources","Baseline Start","Baseline Finish","Total Slack (day)","Critical","Summary","Milestone","Ghi chú"]
        ws.append(headers)
        for t in tasks:
            delta=progress_delta(t["planned_progress"],t["actual_progress"])
            delay_days=calculate_delay_days(t["end_date"],t["actual_progress"],t["actual_finish_date"] or "")
            ws.append([t["source_task_id"],t["source_uid"],t["wbs"],t["name"],t["start_date"],t["end_date"],t["duration"],t["planned_progress"],t["actual_progress"],progress_evaluation(t["planned_progress"],t["actual_progress"]),delta,t["status"],delay_days,t["actual_update_date"],t["actual_finish_date"],t["predecessor"],t["resource_names"],t["baseline_start"],t["baseline_finish"],t["total_slack"],t["critical"],t["is_summary"],t["is_milestone"],t["note"]])
        ws.freeze_panes="A2"
        for col in ws.columns:
            width=min(55,max(11,max(len(str(cell.value or "")) for cell in col)+2)); ws.column_dimensions[col[0].column_letter].width=width
        wb.save(path); QMessageBox.information(self,"Hoàn tất",f"Đã xuất Excel:\n{path}")

    def import_excel(self):
        if load_workbook is None: QMessageBox.warning(self,"Thiếu thư viện","Cài openpyxl: pip install openpyxl"); return
        if not self.current_project_id: QMessageBox.information(self,"Chưa có dự án","Hãy chọn dự án trước."); return
        path,_=QFileDialog.getOpenFileName(self,"Nhập tiến độ Excel","","Excel (*.xlsx)")
        if not path: return
        try:
            wb=load_workbook(path,data_only=True); ws=wb.active; rows=list(ws.iter_rows(values_only=True))
            if len(rows)<2: raise ValueError("File không có dữ liệu")
            headers=[str(x or "").strip().lower() for x in rows[0]]
            aliases={"wbs":["wbs"],"name":["công việc","cong viec","task","name"],"responsible":["phụ trách","phu trach","responsible","owner"],"start_date":["bắt đầu","bat dau","start","start date"],"end_date":["kết thúc","ket thuc","finish","end date"],"planned_progress":["kh %","planned %","plan %"],"actual_progress":["tt %","actual %","progress %"],"predecessor":["công việc trước","cong viec truoc","predecessor"],"note":["ghi chú","ghi chu","note"]}
            idx={k:next((headers.index(n) for n in names if n in headers),None) for k,names in aliases.items()}
            if idx["name"] is None or idx["start_date"] is None or idx["end_date"] is None: raise ValueError("Cần tối thiểu: Công việc, Bắt đầu, Kết thúc")
            def cell(row,key,default=""):
                i=idx.get(key); return default if i is None or i>=len(row) or row[i] is None else row[i]
            def dstr(v):
                if isinstance(v,datetime): return v.date().strftime(DATE_FMT)
                if isinstance(v,date): return v.strftime(DATE_FMT)
                s=str(v).strip()
                for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%m/%d/%Y"):
                    try: return datetime.strptime(s,fmt).strftime(DATE_FMT)
                    except ValueError: pass
                raise ValueError(f"Không đọc được ngày: {v}")
            count=0
            for row in rows[1:]:
                name=str(cell(row,"name","")).strip()
                if not name: continue
                ss,es=dstr(cell(row,"start_date")),dstr(cell(row,"end_date")); s=datetime.strptime(ss,DATE_FMT).date(); e=datetime.strptime(es,DATE_FMT).date(); planned=int(float(cell(row,"planned_progress",0) or 0)); actual=int(float(cell(row,"actual_progress",0) or 0))
                d=dict(project_id=self.current_project_id,wbs=str(cell(row,"wbs","")),name=name,responsible=str(cell(row,"responsible","")),start_date=ss,end_date=es,duration=max(1,(e-s).days+1),planned_progress=max(0,min(100,planned)),actual_progress=max(0,min(100,actual)),status=TaskDialog.calc_status(ss,es,planned,actual),predecessor=str(cell(row,"predecessor","")),note=str(cell(row,"note","")),source_type="manual",source_uid=None,source_task_id=None,outline_level=1,is_summary=0,is_milestone=0,critical=0,total_slack=0,resource_names="",baseline_start="",baseline_finish="")
                self.db.add_task(d); count+=1
            self.refresh_tasks(); QMessageBox.information(self,"Hoàn tất",f"Đã nhập {count} công việc.")
        except Exception as exc: QMessageBox.critical(self,"Lỗi nhập Excel",str(exc))


def main():
    app=QApplication(sys.argv); app.setStyle("Fusion"); w=MainWindow(); w.show(); sys.exit(app.exec())

if __name__=="__main__": main()
